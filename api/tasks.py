from celery import shared_task
from datetime import timedelta
from time import monotonic_ns, sleep
from ui.models import Task
import requests
import os
from testrail import *
import re
from shutil import unpack_archive
from django.conf import settings
import openpyxl

TESTRAIL_BASE_URL = os.environ.get('TESTRAIL_URL', 'https://tele2se.testrail.net/')
TESTRAIL = APIClient(TESTRAIL_BASE_URL)
ATTACHMENT_URL_PREFIX = f'{TESTRAIL_BASE_URL}index.php?/attachments/get/'
TEST_CASE_URL_PREFIX = f'{TESTRAIL_BASE_URL}index.php?/cases/view/'

PROJECT_ID      = None      # Evaluate Testrail Baltics option 2
TYPE_ID         = 9         # Regression
PRIORITY_ID     = 2         # Medium
SUITE_ID        = None      # Evaluate Testrail Baltics option 2 > Regression 1312
TEMPLATE_ID     = 2         # Test Case (Steps)
TOP_SECTION     = ''
TOP_SECTION_ID  = None

CUSTOM_FIELD_BUSINESS_CRITICAL    = "custom_businesscritical"
CUSTOM_FIELD_STATUS_FOR_LVI       = "custom_statusforlvi"
CUSTOM_FIELD_STATUS_FOR_LTH       = "custom_statusforlth"
CUSTOM_FIELD_STATUS_FOR_ETI       = "custom_statusforeti"
CUSTOM_FIELD_STATUS_FOR_SVE       = "custom_statusforsve"
CUSTOM_FIELD_REQUIREMENT_COVERAGE = "custom_requirementcoverage"
CUSTOM_FIELD_MAPPING_FIELD_SAF    = "custom_mappingfieldinsaf"
CUSTOM_FIELD_QC_REF               = "custom_qcidreference"

ALL_TESTRAIL_SECTIONS = []  # Local cache for Testrail sections

BUSINESS_CRITICALMAPPING = {
    "Tier 1":           1,
    "Tier 2":           2,
    "Tier 3":           3,
    "xKeyscenarioT1":   4,
}

STATUS_FOR_COUNTRY_MAPPING = {
    "10-Not for Automation":    1,
    "1-Draft":                  2,
    "2-Open":                   3,
    "3-Designed":               4,
    "4-Approved":               5,
    "5- Must Run in RT":        6,
    "6-Rejected":               7,
    "7-Cancelled":              8,
    "8-Ready For Automation":   9,
    "9-Broken":                 10,
}

def make_api_get_request(uri):
    too_many_requests = False
    while not too_many_requests:
        try:
            response = TESTRAIL.send_get(uri)
            return response
        except APIError as error:
            error_string = str(error)
            if 'Retry after' in error_string:
                # Parse retry after x seconds
                retry_after = error_string.split('Retry after ')[1]
                retry_after = retry_after.split(' ', 1)[0]
                retry_after = int(retry_after)
                print('Pause for %x seconds' % retry_after)
                sleep(retry_after)
                too_many_requests = True
            elif '504' in error_string or '502' in error_string:
                print('%s error: Pause for 10 seconds' % error_string)
                sleep(10)
                too_many_requests = True
            else:
                raise Exception('Unexpected API Error: %s' % error_string)

def make_api_post_request(uri, body):
    too_many_requests = False
    while not too_many_requests:
        try:
            response = TESTRAIL.send_post(uri, body)
            return response
        except APIError as error:
            error_string = str(error)
            if 'Retry after' in error_string:
                # Parse retry after x seconds
                retry_after = error_string.split('Retry after ')[1]
                retry_after = retry_after.split(' ', 1)[0]
                retry_after = int(retry_after)
                print('Pause for %x seconds' % retry_after)
                sleep(retry_after)
                too_many_requests = True
            elif '504' in error_string or '502' in error_string:
                print('%s error: Pause for 10 seconds' % error_string)
                sleep(10)
                too_many_requests = True
            else:
                raise Exception('Unexpected API Error: %s' % error_string)

def get_all_sections(project_id, suite_id):
    uri = 'get_sections/' + str(project_id) + '&suite_id=' + str(suite_id)
    sections = []
    section_list = make_api_get_request(uri)
    while section_list['_links']['next']:
        sections.extend(section_list['sections'])
        uri = section_list['_links']['next'].replace('/api/v2/', '')
        section_list = make_api_get_request(uri)
    else:
        sections.extend(section_list['sections'])
    return sections

def create_section(project_id, suite_id, parent_id, name):
    uri = 'add_section/' + str(project_id)
    if parent_id is None:
        request_body = {
            'suite_id': suite_id,
            'name': name
        }
    else:
        request_body = {
            'suite_id': suite_id,
            'parent_id': parent_id,
            'name': name
        }

    section = make_api_post_request(uri, request_body)
    return section['id']

def create_test_case(section_id, test_case):
    uri = 'add_case/' + str(section_id)
    print(test_case)
    case = make_api_post_request(uri, test_case)
    return case['id']

def update_test_case(case_id, test_case):
    uri = 'update_case/' + str(case_id)

    case = make_api_post_request(uri, test_case)
    return case['id']

def upload_attachment(case_id, file_path):
    too_many_requests = False
    uri = 'add_attachment_to_case/' + str(case_id)
    while not too_many_requests:
        try:
            response = TESTRAIL.send_post(uri, file_path)
            return response['attachment_id']
        except APIError as error:
            error_string = str(error)
            if 'Retry after' in error_string:
                # Parse retry after x seconds
                retry_after = error_string.split('Retry after ')[1]
                retry_after = retry_after.split(' ', 1)[0]
                retry_after = int(retry_after)
                print('Pause for %x seconds' % retry_after)
                sleep(retry_after)
                too_many_requests = True
            else:
                raise Exception('Unexpected API Error: %s' % error_string)

def check_and_create_section_hier(parent_section_name, section_name):
    global ALL_TESTRAIL_SECTIONS
    global TOP_SECTION_ID
    global TOP_SECTION
    
    top_section_exists = False
    top_section_id = -1

    for section in ALL_TESTRAIL_SECTIONS:
        if TOP_SECTION_ID == section['id']:
            top_section_id = section['id']
            top_section_exists = True
            break
    
    if not top_section_exists:
        all_sections = get_all_sections(PROJECT_ID, SUITE_ID)
        for section in all_sections:
            if TOP_SECTION == section['name']:
                top_section_id = section['id']
                ALL_TESTRAIL_SECTIONS.append({
                    'name': section['name'],
                    'id': section['id'],
                    'parent_id': None,
                })
                top_section_exists = True
                break
        if not top_section_exists:
            top_section_id = create_section(PROJECT_ID, SUITE_ID, None, TOP_SECTION)
            ALL_TESTRAIL_SECTIONS.append({
                'name': TOP_SECTION,
                'id': top_section_id,
                'parent_id': None,
            })

    if parent_section_name == '!Categories':
        return top_section_id
    
    if parent_section_name == TOP_SECTION:
        parent_section_id = -1
        section_id = -1
        for section in ALL_TESTRAIL_SECTIONS:
            if parent_section_name == section['name'] and TOP_SECTION == section['name']:
                parent_section_id = section['id']
                break
        if parent_section_id == -1:
            all_sections = get_all_sections(PROJECT_ID, SUITE_ID)
            for section in all_sections:
                if parent_section_name == section['name'] and TOP_SECTION == section['name']:
                    parent_section_id = section['id']
                    ALL_TESTRAIL_SECTIONS.append({
                        'name': section['name'],
                        'id': section['id'],
                        'parent_id': None,
                    })
                    break
            if parent_section_id == -1:
                parent_section_id = create_section(PROJECT_ID, SUITE_ID, None, parent_section_name)
                ALL_TESTRAIL_SECTIONS.append({
                    'name': parent_section_name,
                    'id': parent_section_id,
                    'parent_id': None,
                })
        
        for section in ALL_TESTRAIL_SECTIONS:
            if section_name == section['name'] and parent_section_id == section['parent_id']:
                section_id = section['id']
                break
        if section_id == -1:
            all_sections = get_all_sections(PROJECT_ID, SUITE_ID)
            for section in all_sections:
                if section_name == section['name'] and parent_section_id == section['parent_id']:
                    section_id = section['id']
                    ALL_TESTRAIL_SECTIONS.append({
                        'name': section['name'],
                        'id': section['id'],
                        'parent_id': section['parent_id'],
                    })
                    break
            if section_id == -1:
                section_id = create_section(PROJECT_ID, SUITE_ID, parent_section_id, section_name)
                ALL_TESTRAIL_SECTIONS.append({
                    'name': section_name,
                    'id': section_id,
                    'parent_id': parent_section_id,
                })
        return section_id
    else:
        parent_section_id = -1
        section_id = -1
        for section in ALL_TESTRAIL_SECTIONS:
            if parent_section_name == section['name'] and top_section_id == section['parent_id']:
                parent_section_id = section['id']
                break
        if parent_section_id == -1:
            all_sections = get_all_sections(PROJECT_ID, SUITE_ID)
            for section in all_sections:
                if parent_section_name == section['name'] and top_section_id == section['parent_id']:
                    parent_section_id = section['id']
                    ALL_TESTRAIL_SECTIONS.append({
                        'name': section['name'],
                        'id': section['id'],
                        'parent_id': section['parent_id'],
                    })
                    break
            if parent_section_id == -1:
                parent_section_id = create_section(PROJECT_ID, SUITE_ID, top_section_id, parent_section_name)
                ALL_TESTRAIL_SECTIONS.append({
                    'name': parent_section_name,
                    'id': parent_section_id,
                    'parent_id': top_section_id,
                })

        for section in ALL_TESTRAIL_SECTIONS:
            if section_name == section['name'] and parent_section_id == section['parent_id']:
                section_id = section['id']
                break

        if section_id == -1:  
            all_sections = get_all_sections(PROJECT_ID, SUITE_ID)
            for section in all_sections:
                if section_name == section['name'] and parent_section_id == section['parent_id']:
                    section_id = section['id']
                    ALL_TESTRAIL_SECTIONS.append({
                        'name': section['name'],
                        'id': section['id'],
                        'parent_id': section['parent_id'],
                    })
                    break
            if section_id == -1:
                section_id = create_section(PROJECT_ID, SUITE_ID, parent_section_id, section_name)
                ALL_TESTRAIL_SECTIONS.append({
                    'name': section_name,
                    'id': section_id,
                    'parent_id': parent_section_id,
                })
        return section_id

@shared_task(bind=True)
def import_task(self, task_id, username, password, auth_header):
    global TESTRAIL
    global PROJECT_ID
    global TYPE_ID
    global PRIORITY_ID
    global SUITE_ID
    global TEMPLATE_ID
    global TOP_SECTION
    global TOP_SECTION_ID

    imported_cases = 0

    hostname = os.environ.get('CURRENT_HOST', 'http://localhost:8000')
    TESTRAIL.user = username
    TESTRAIL.password = password

    current_task = Task.objects.get(id=task_id)
    current_user = current_task.user

    PROJECT_ID = current_task.project_id
    SUITE_ID = current_task.suite_id
    TOP_SECTION_ID = current_task.section_id
    TOP_SECTION = current_task.section_name

    qc_id_file = settings.MEDIA_ROOT + '/' + str(current_task.id_file_name)
    qc_dump_file = settings.MEDIA_ROOT + '/' + str(current_task.steps_file_name)
    attachment_file = settings.MEDIA_ROOT + '/' + str(current_task.attachment_file_name)

    current_task.status = "In Progress"
    current_task.save()
    requests.post(f'{hostname}/api/create_event/', json={"user": current_user}, headers={"Authorization": auth_header}, verify=False)

    start = monotonic_ns()
    unpack_archive(attachment_file, settings.MEDIA_ROOT + '/' + current_task.user + '/' + current_task.session_id)

    # loads the test case ids extracted from qc
    try:
        wb_id = openpyxl.load_workbook(qc_id_file, data_only=True)
        ws_id = wb_id.worksheets[0]
        max_row_id = ws_id.max_row
    except OSError:
        end = monotonic_ns()
        current_task.status_message = "Can't load QC Ids file. Exit."
        current_task.status = "Failed"
        current_task.elapsed_time = timedelta(milliseconds=int((end-start)/1000000))
        current_task.save()
        requests.post(f'{hostname}/api/create_event/', json={"user": current_user}, headers={"Authorization": auth_header}, verify=False)
        return
    
    qc_ids = []
    qc_id_with_req = {}
    for i in range(2, max_row_id + 1):
        qc_ids.append(ws_id.cell(row=i, column=1).value)
        qc_id_with_req[str(ws_id.cell(row=i, column=1).value)] = str(ws_id.cell(row=i, column=3).value)
    wb_id.close()

    # starts processing the dump file
    try:
        wb = openpyxl.load_workbook(qc_dump_file, data_only=True)
        ws = wb.worksheets[0]
        max_row = ws.max_row
    except OSError:
        end = monotonic_ns()
        current_task.status_message = "Can't load QC Test Steps file. Exit."
        current_task.status = "Failed"
        current_task.elapsed_time = timedelta(milliseconds=int((end-start)/1000000))
        current_task.save()
        requests.post(f'{hostname}/api/create_event/', json={"user": current_user}, headers={"Authorization": auth_header}, verify=False)
        return

    # Update total count excluding processed cases
    if current_task.retry_import:
        current_task.total_cases = len(qc_ids)
    else:
        unprocessed_test_case_ids = []
        for i in range(2, max_row + 1):
            if ws.cell(row=i, column=32).value is None:
                if ws.cell(row=i, column=3).value not in unprocessed_test_case_ids:
                    unprocessed_test_case_ids.append(ws.cell(row=i, column=3).value)
        current_task.total_cases = len(list(set(qc_ids) & set(unprocessed_test_case_ids)))
    current_task.save()
    requests.post(f'{hostname}/api/create_event/', json={"user": current_user}, headers={"Authorization": auth_header}, verify=False)

    # starts reading test cases and pushing to Jira
    start_step = 1
    test_step_desc = ''
    test_case = None
    testrail_case_id = None
    attachment_desc = ''
    try:
        for i in range(2, max_row + 1):
            # Check if Test case already imported or user retry
            if not ws.cell(row=i, column=32).value is None and not current_task.retry_import:
                continue

            # Check if Test Case Id is within input range
            test_case_id = ws.cell(row=i, column=3).value
            if test_case_id not in qc_ids:
                continue

            # First Step
            if start_step == 1:
                test_case_parent_dir = str(ws.cell(row=i, column=1).value)
                test_case_dir = str(ws.cell(row=i, column=2).value)
                test_case_title = str(ws.cell(row=i, column=4).value).replace('_x000D_', '').replace('\n', ' ') if not ws.cell(row=i, column=4).value is None else ''
                test_case_precond = str(ws.cell(row=i, column=5).value).replace('_x000D_', '') if not ws.cell(row=i, column=5).value is None else ''
                test_case_comments = str(ws.cell(row=i, column=6).value).replace('_x000D_', '') if not ws.cell(row=i, column=6).value is None else ''
                test_case_precond = test_case_precond + '\nComments: ' + test_case_comments if test_case_comments else test_case_precond
                steps_count = ws.cell(row=i, column=7).value
                saf_field = str(ws.cell(row=i, column=21).value) if not ws.cell(row=i, column=21).value is None else ''
                business_critical = BUSINESS_CRITICALMAPPING.get(str(ws.cell(row=i, column=20).value), '')  if not ws.cell(row=i, column=20).value is None else ''
                status_for_sve = STATUS_FOR_COUNTRY_MAPPING.get(str(ws.cell(row=i, column=22).value), '')  if not ws.cell(row=i, column=22).value is None else ''
                status_for_eti = STATUS_FOR_COUNTRY_MAPPING.get(str(ws.cell(row=i, column=23).value), '')  if not ws.cell(row=i, column=23).value is None else ''
                status_for_lth = STATUS_FOR_COUNTRY_MAPPING.get(str(ws.cell(row=i, column=24).value), '')  if not ws.cell(row=i, column=24).value is None else ''
                status_for_lvi = STATUS_FOR_COUNTRY_MAPPING.get(str(ws.cell(row=i, column=25).value), '')  if not ws.cell(row=i, column=25).value is None else ''
                requirement_coverage = qc_id_with_req[str(test_case_id)]

                # Check and Create section hierarchy
                section_id = check_and_create_section_hier(test_case_parent_dir, test_case_dir)
                # Create Test Case
                test_case = {
                    'title': 'QC-' + str(test_case_id) + ' | ' + test_case_title,
                    'section_id': section_id,
                    'template_id': TEMPLATE_ID,
                    'type_id': TYPE_ID,
                    'priority_id': PRIORITY_ID,
                    'suite_id': SUITE_ID,
                    'custom_preconds': test_case_precond,
                    CUSTOM_FIELD_QC_REF: str(test_case_id),
                    CUSTOM_FIELD_BUSINESS_CRITICAL: business_critical,
                    CUSTOM_FIELD_STATUS_FOR_LVI: status_for_lvi,
                    CUSTOM_FIELD_STATUS_FOR_LTH: status_for_lth,
                    CUSTOM_FIELD_STATUS_FOR_ETI: status_for_eti,
                    CUSTOM_FIELD_STATUS_FOR_SVE: status_for_sve,
                    CUSTOM_FIELD_REQUIREMENT_COVERAGE: requirement_coverage,
                    CUSTOM_FIELD_MAPPING_FIELD_SAF: saf_field,
                    'custom_steps_separated': []
                }
                testrail_case_id = create_test_case(test_case['section_id'], test_case)

            # Other steps
            test_step_id = str(ws.cell(row=i, column=10).value)
            next_step_id = str(ws.cell(row=i+1, column=10).value) if not ws.cell(row=i+1, column=10).value is None else ''
            test_step_attachment_path = ws.cell(row=i, column=17).value if not ws.cell(row=i, column=17).value is None else ''
            test_step_attachment_name = re.sub(r'attachments.*\\', '', ws.cell(row=i, column=17).value) if not ws.cell(row=i, column=17).value is None else ''
            if test_step_attachment_name != '':
                test_step_attachment_path = settings.MEDIA_ROOT + '/' + current_task.user + '/' + current_task.session_id + '/' + test_step_attachment_path.replace('\\', '/')
                attachment_id = upload_attachment(testrail_case_id, test_step_attachment_path)
                attachment_desc += f'\n[{test_step_attachment_name}]({ATTACHMENT_URL_PREFIX}{attachment_id})'
            if test_step_id != next_step_id:
                test_step_name = str(ws.cell(row=i, column=11).value)
                test_step_desc = str(ws.cell(row=i, column=13).value).replace('_x000D_', '') if not ws.cell(row=i, column=13).value is None else ''
                test_step_desc += attachment_desc
                attachment_desc = ''
                test_step_exp_result = str(ws.cell(row=i, column=14).value).replace('_x000D_', '') if not ws.cell(row=i, column=14).value is None else ''
                test_case['custom_steps_separated'].append({
                    'content': test_step_desc,
                    'expected': test_step_exp_result
                })
            ws.cell(row=i, column=32).value = str(testrail_case_id)
            ws.cell(row=i, column=33).value = TEST_CASE_URL_PREFIX + str(testrail_case_id)

            # Last Step
            if start_step == steps_count:
                if test_step_id != next_step_id:
                    testrail_case_id = update_test_case(testrail_case_id, test_case)
                    imported_cases += 1
                    current_task.imported_cases = imported_cases
                    current_task.save()
                    requests.post(f'{hostname}/api/create_event/', json={"user": current_user}, headers={"Authorization": auth_header}, verify=False)
                    test_case = None
                    testrail_case_id = None
                    start_step = 1
                    continue
            if test_step_id != next_step_id:
                start_step += 1

        wb.save(qc_dump_file)
        wb.close()
        end = monotonic_ns()
        current_task.status_message = "Test cases imported. Successfully updated dump file with Testrail Ids. Exit."
        current_task.status = "Complete"
        current_task.elapsed_time = timedelta(milliseconds=int((end-start)/1000000))
        current_task.save()
        requests.post(f'{hostname}/api/create_event/', json={"user": current_user}, headers={"Authorization": auth_header}, verify=False)
    except Exception as e:
        wb.save(qc_dump_file)
        wb.close()
        end = monotonic_ns()
        current_task.status_message = "Error: " + str(e) + ". Exit."
        current_task.status = "Failed"
        current_task.elapsed_time = timedelta(milliseconds=int((end-start)/1000000))
        current_task.save()
        requests.post(f'{hostname}/api/create_event/', json={"user": current_user}, headers={"Authorization": auth_header}, verify=False)